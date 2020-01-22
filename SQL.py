import mysql.connector

def SQLConnect():
    conn = mysql.connector.connect(host = "localhost", user = "root",password = "checco")
    #conn = mysql.connector.connect(host = 'localhost', user = 'root', password = 'password')
    #conn = mysql.connector.connect(host = 'localhost', user = 'root', password = 'sole1997')
    
    if checkDb(conn) is True:
        createDb(conn)
        queryUse(conn)
        executeQueries('diagn_title', 'diagn',conn)
        executeQueries('diffsydiw', 'diff',conn)
        executeQueries('symptoms2', 'sym',conn)
    else:
        print("Database esistente!")
        queryUse(conn)
    
    return conn


def checkDb(conn):
    #Funzione che controlla se il database è stato già creato
    cursor = conn.cursor()
    
    queryCheck = """SHOW DATABASES LIKE 'medical';"""
    cursor.execute(queryCheck)
    check = cursor.fetchall()
    
    if not check:
        return True
    else:
        return False


def createDb(conn):
    import pymysql as sql    
    sql.install_as_MySQLdb()
    import MySQLdb as mysql
    
    cursor = conn.cursor()
    query1 = """CREATE DATABASE medical; """
    query2 = """USE medical; """
    
    try:
         cursor.execute(query1)
         cursor.execute(query2)
         conn.commit()
    except sql.ProgrammingError:
         pass
    
    cursor.close()
    

def queryUse(conn):
    cursor = conn.cursor()
    query = """USE medical;"""
    cursor.execute(query)
    conn.commit()
    

def executeQueries(filename,tablename,conn):
    import pymysql as sql    
    sql.install_as_MySQLdb()
    import MySQLdb as mysql
    import openpyxl as xl
    
    cursor = conn.cursor()  
    queryDrop = """DROP TABLE IF EXISTS """+ tablename
    
    if filename == 'diagn_title':
        
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['id']
    
        #Query per la creazione della tabella del file diagn_title.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    id VARCHAR(30),
                    title text,
                    cat CHAR(9));"""
 
       #Query per l'inserimento dei dati nella tabella
        insertQuery = """INSERT INTO """+ tablename +"""(
                     id,title,cat) VALUES(%s,%s,%s);"""
        
    elif filename == 'diffsydiw':
        
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['syd']
        
        #Query per la creazione della tabella del file diffsydiw.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    syd CHAR(4),
                    did CHAR(4));"""
        
        #Query per l'inserimento dei dati nella tabella
        insertQuery = """INSERT INTO """+ tablename +"""(
                     syd,did) VALUES(%s,%s);"""
        
    elif filename == 'symptoms2':    
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['_id']
        
        #Query per la creazione della tabella del file symptoms2.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    _id CHAR(4),
                    name text,
                    cat CHAR(9));"""
        
        #Query per l'inserimento dei dati nella tabella
        insertQuery = """INSERT INTO """+ tablename +"""(
                     _id ,name,cat) VALUES(%s,%s,%s);"""
    try:
        cursor.execute(queryDrop)
        cursor.execute(queryCreate)
        conn.commit()
    except sql.ProgrammingError:
        pass
    
    for row in table.rows:
        
        if filename == 'symptoms2':
             _id = row[1].value
             name = row[3].value
             cat = row[9].value
             values = (_id,name,cat)
             #Esecuzione query
             cursor.execute(insertQuery,values)
        elif filename == 'diffsydiw':
            syd = row[0].value
            did = row[1].value
            values = (syd,did)
            #Esecuzione query
            cursor.execute(insertQuery,values)
        elif filename == 'diagn_title':
            id = row[0].value
            title = row[1].value
            cat = row[2].value
            values = (id,title,cat)
            #Esecuzione query
            cursor.execute(insertQuery,values)
    cursor.close()      
    #Commit della transazione
    conn.commit()
    

def checkSymWithErr(conn,value):
    #Funzione che verifica se un determinato sintomo inserito è corretto
    cursor = conn.cursor(buffered = True)
    querySym = """SELECT _id from sym where name='""" + value +"""';""" 
    cursor.execute(querySym)
    check = cursor.fetchall()
    if not check:
        return False
    else:
        return True
        
    
def searchDiagn(conn,list):
    #Funzione che, data una lista di sintomi, restituisce l'intera lista di diagnosi ad essi corrispondenti
    import numpy as np
    
    cursor = conn.cursor(buffered = True)
    idDiseases = []
    idSymptoms = []
    
    for i in list:
        #Query che preleva l'id corrispondenti ai sintomi
        querySym = """SELECT DISTINCT _id from sym where name='""" + i +"""';""" 
        cursor.execute(querySym)
        idSymptoms.append(cursor.fetchall()) 

    for i in idSymptoms:
        i = np.asarray(i)
        #Query che restituisce l'id delle malattie collegate all'id dei sintomi(colonna did nella tabella)
        queryDid = """SELECT did from diff where syd='"""+ i[0][0] +"""';"""
        cursor.execute(queryDid)
        idDiseases.append(cursor.fetchall())
        
    return idDiseases
    
        
def searchSymCategories(conn,list):
    #Funzione che restituisce le categorie per ogni sintomo inserito
    
    cursor = conn.cursor(buffered = True)
    categories = []
    
    for i in list:
        #Query che preleva le categorie corrispondenti ai sintomi
        querySym = """SELECT cat from sym where name='""" + i +"""';""" 
        cursor.execute(querySym)
        categories.append(cursor.fetchall())
    
    return categories
   
    
def checkDiagnName(cursor, idDiagn):
    #Funzione che restituisce il nome di una diagnosi dato il suo id
    queryName = """SELECT title from diagn where id ='"""+idDiagn[0]+"""';"""
    cursor.execute(queryName) 
    cursor = cursor.fetchall()
    cursor = cursor[0][0]
    if "_" in cursor:
        pos = cursor.index("_")
        cursor = cursor[0:pos]
    
    return cursor

def searchCatDiagn(conn,diagn):
    #funzione che restituisce le categoria di una diagnosi
    query = """SELECT cat FROM diagn WHERE id = """+ diagn + """;"""
    cursor = conn.cursor(buffered = True)
    cursor.execute(query)
    cursor = cursor.fetchall()
    return cursor[0]

def closeConn(conn):
    conn.close()
    
    